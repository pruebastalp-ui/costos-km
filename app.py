from collections import defaultdict
import copy
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash,  send_file
from config import SECRET_KEY
from config import MYSQL_HOST
from db import close_db, query_all, query_one, execute, executemany, commit, rollback
from services.calculator import calculate_results
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.teardown_appcontext(close_db)

# agregue desde aca
EMPRESA_ORDER = [
    "parque_movil",
    "km_productivo_anual_veh",
    "ipk",
    "pasajeros_transportados",
    "km_anual_empresa",
    "recaudacion_venta_boletos",
    "km_improductivo_pct",
    "vida_util_vehiculo_km",
    "velocidad_comercial",
    "recaudacion_media_km",
    "antiguedad_media_parque",
    "ingreso_medio_pasajero",
]

FORMAT_0_DEC = {
    "km_anual_empresa",
    "pasajeros_transportados",
}

FORMAT_4_DEC = {
    "ipk",
}

# hasta aca


@app.context_processor
def inject_globals():
    return dict(VERSION_TIPO="CMV")


def load_groups():
    return query_all("SELECT * FROM grupos_tarifarios ORDER BY orden")


def load_parameters_with_values(scenario_id=None):
    rows = query_all(
        """
        SELECT p.id AS parametro_id,
               p.codigo,
               p.modulo,
               p.descripcion,
               p.unidad,
               g.codigo AS grupo_codigo,
               g.nombre AS grupo_nombre,
               b.valor AS valor_base,
               v.valor AS valor_escenario
        FROM definiciones_parametros p
        CROSS JOIN grupos_tarifarios g
        LEFT JOIN valores_base b
               ON b.parametro_id = p.id
              AND b.grupo_id = g.id
        LEFT JOIN valores_escenario v
               ON v.parametro_id = p.id
              AND v.grupo_id = g.id
              AND v.escenario_id = %s
        ORDER BY p.modulo, p.id, g.orden
        """,
        (scenario_id if scenario_id else -1,)
    )

    grouped = defaultdict(lambda: defaultdict(dict))
    merged = defaultdict(dict)

    for r in rows:
        modulo = r["modulo"]
        codigo = r["codigo"]
        grupo_codigo = r["grupo_codigo"]

        grouped[modulo][codigo]["code"] = codigo
        grouped[modulo][codigo]["label"] = r["descripcion"]
        grouped[modulo][codigo]["unit"] = r["unidad"]

        if codigo in FORMAT_0_DEC:
            grouped[modulo][codigo]["display_decimals"] = 0
        elif codigo in FORMAT_4_DEC:
            grouped[modulo][codigo]["display_decimals"] = 4
        else:
            grouped[modulo][codigo]["display_decimals"] = 2

        grouped[modulo][codigo].setdefault("values", {})
        grouped[modulo][codigo].setdefault("base_values", {})
        grouped[modulo][codigo].setdefault("changed", {})

        base_value = float(r["valor_base"]) if r["valor_base"] is not None else 0.0

        if r["valor_escenario"] is not None:
            effective = float(r["valor_escenario"])
        elif r["valor_base"] is not None:
            effective = float(r["valor_base"])
        else:
            effective = 0.0

        grouped[modulo][codigo]["values"][grupo_codigo] = effective
        grouped[modulo][codigo]["base_values"][grupo_codigo] = base_value
        grouped[modulo][codigo]["changed"][grupo_codigo] = abs(effective - base_value) > 1e-12

        merged[codigo][grupo_codigo] = effective

    # Reordenar solo el módulo Empresa
    if "empresa" in grouped:
        empresa_ordenada = {}
        for code in EMPRESA_ORDER:
            if code in grouped["empresa"]:
                empresa_ordenada[code] = grouped["empresa"][code]

        for code, item in grouped["empresa"].items():
            if code not in empresa_ordenada:
                empresa_ordenada[code] = item

        grouped["empresa"] = empresa_ordenada

    return grouped, merged

def load_base_params():
    rows = query_all(
        """
        SELECT p.codigo, g.codigo AS grupo_codigo, b.valor
        FROM valores_base b
        JOIN definiciones_parametros p ON p.id = b.parametro_id
        JOIN grupos_tarifarios g ON g.id = b.grupo_id
        """
    )
    out = defaultdict(dict)
    for r in rows:
        out[r["codigo"]][r["grupo_codigo"]] = float(r["valor"])
    return out


def load_baseline_results():
    rows = query_all(
        """
        SELECT d.codigo, g.codigo AS grupo_codigo, r.valor
        FROM resultados_base r
        JOIN definiciones_resultados d ON d.id = r.resultado_id
        JOIN grupos_tarifarios g ON g.id = r.grupo_id
        """
    )
    out = defaultdict(dict)
    for r in rows:
        out[r["codigo"]][r["grupo_codigo"]] = float(r["valor"])
    return out


def persist_results(scenario_id, results):
    defs = query_all("SELECT id, codigo FROM definiciones_resultados")
    result_id_map = {r["codigo"]: r["id"] for r in defs}

    groups = query_all("SELECT id, codigo FROM grupos_tarifarios")
    group_id_map = {g["codigo"]: g["id"] for g in groups}

    rows = []
    for code, per_group in results.items():
        if code not in result_id_map:
            continue
        for group_code, value in per_group.items():
            if group_code not in group_id_map:
                continue
            rows.append(
                (
                    scenario_id,
                    result_id_map[code],
                    group_id_map[group_code],
                    float(value),
                )
            )

    if rows:
        executemany(
            """
            INSERT INTO resultados_escenario (escenario_id, resultado_id, grupo_id, valor)
            VALUES (%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                valor = VALUES(valor),
                calculado_en = NOW()
            """,
            rows
        )


def build_impact_summary(ordered_results, groups, top_n=5):
    impact_by_group = {}

    for g in groups:
        group_code = g["codigo"]
        rows = []

        for code, item in ordered_results:
            value = item["values"].get(group_code, 0.0)
            base_value = item["base_values"].get(group_code, 0.0)
            delta = value - base_value

            if abs(delta) <= 1e-12:
                continue

            pct = 0.0
            if abs(base_value) > 1e-12:
                pct = (delta / base_value) * 100.0

            rows.append(
                {
                    "code": code,
                    "label": item["label"],
                    "value": value,
                    "base_value": base_value,
                    "delta": delta,
                    "pct": pct,
                    "abs_delta": abs(delta),
                }
            )

        rows.sort(key=lambda x: x["abs_delta"], reverse=True)
        impact_by_group[group_code] = rows[:top_n]

    return impact_by_group


def calculate_results_isolated(base_params, merged_params, baseline_results, groups):
    """
    Ejecuta el cálculo varias veces sobre copias completas de la estructura,
    y conserva en cada corrida solamente el grupo objetivo.
    Esto evita contaminación entre grupos sin romper fórmulas que esperan
    todas las claves tarifarias presentes.
    """
    final_results = defaultdict(dict)

    for g in groups:
        group_code = g["codigo"]

        base_params_copy = copy.deepcopy(base_params)
        merged_params_copy = copy.deepcopy(merged_params)
        baseline_results_copy = copy.deepcopy(baseline_results)

        calc_all = calculate_results(
            base_params_copy,
            merged_params_copy,
            baseline_results_copy
        )

        for result_code, per_group in calc_all.items():
            if group_code in per_group:
                final_results[result_code][group_code] = float(per_group[group_code])

    return final_results


@app.route("/")
def index():
    escenarios = query_all(
        """
        SELECT
            e.*,

            (
                SELECT re.valor
                FROM resultados_escenario re
                JOIN definiciones_resultados dr
                  ON dr.id = re.resultado_id
                JOIN grupos_tarifarios gt
                  ON gt.id = re.grupo_id
                WHERE re.escenario_id = e.id
                  AND dr.codigo = 'costo_total_km_sin_iva'
                  AND gt.codigo = e.grupo_codigo
                LIMIT 1
            ) AS costo_total_km_sin_iva_grupo,

            (
                SELECT re.valor
                FROM resultados_escenario re
                JOIN definiciones_resultados dr
                  ON dr.id = re.resultado_id
                JOIN grupos_tarifarios gt
                  ON gt.id = re.grupo_id
                WHERE re.escenario_id = e.id
                  AND dr.codigo = 'tarifa_media'
                  AND gt.codigo = e.grupo_codigo
                LIMIT 1
            ) AS tarifa_media_grupo,

            (
                SELECT re.valor
                FROM resultados_escenario re
                JOIN definiciones_resultados dr
                  ON dr.id = re.resultado_id
                JOIN grupos_tarifarios gt
                  ON gt.id = re.grupo_id
                WHERE re.escenario_id = e.id
                  AND dr.codigo = 'ipk'
                  AND gt.codigo = e.grupo_codigo
                LIMIT 1
            ) AS ipk_grupo,

            (
                SELECT re.valor
                FROM resultados_escenario re
                JOIN definiciones_resultados dr
                  ON dr.id = re.resultado_id
                JOIN grupos_tarifarios gt
                  ON gt.id = re.grupo_id
                WHERE re.escenario_id = e.id
                  AND dr.codigo = 'costo_total_sin_imp'
                  AND gt.codigo = e.grupo_codigo
                LIMIT 1
            ) AS costo_total_sin_imp_grupo,

            (
                SELECT rb.valor
                FROM resultados_base rb
                JOIN definiciones_resultados dr
                  ON dr.id = rb.resultado_id
                JOIN grupos_tarifarios gt
                  ON gt.id = rb.grupo_id
                WHERE dr.codigo = 'costo_total_sin_imp'
                  AND gt.codigo = e.grupo_codigo
                LIMIT 1
            ) AS costo_total_sin_imp_base_grupo

        FROM escenarios e
        WHERE e.activo = 1
        ORDER BY e.es_base DESC, e.creado_en DESC, e.id DESC
        """
    )

    groups = load_groups()
    # El valor original antes del cambio para mostrar host del mysql
    #return render_template("index.html", escenarios=escenarios, groups=groups)
    return render_template("index.html",escenarios=escenarios,groups=groups,mysql_host=MYSQL_HOST)
    
@app.route("/escenario/nuevo", methods=["POST"])
def nuevo_escenario():
    nombre = request.form.get("nombre", "").strip()
    descripcion = request.form.get("descripcion", "").strip() or None
    grupo_codigo = request.form.get("grupo_codigo", "").strip() or None

    if not nombre:
        flash("Ingresá un nombre para el escenario.", "error")
        return redirect(url_for("index"))

    if not grupo_codigo:
        flash("Seleccioná un grupo tarifario para el escenario.", "error")
        return redirect(url_for("index"))

    try:
        existe_base = query_one(
            "SELECT id FROM escenarios WHERE es_base = 1 AND activo = 1 LIMIT 1"
        )

        es_base = 0 if existe_base else 1

        execute(
            """
            INSERT INTO escenarios (nombre, descripcion, grupo_codigo, es_base, activo)
            VALUES (%s, %s, %s, %s, 1)
            """,
            (nombre, descripcion, grupo_codigo, es_base)
        )
        commit()

        if es_base:
            flash("Escenario base creado.", "ok")
        else:
            flash("Escenario creado.", "ok")

    except Exception as exc:
        rollback()
        flash(f"No se pudo crear el escenario: {exc}", "error")

    return redirect(url_for("index"))


@app.route("/escenario/<int:scenario_id>/eliminar", methods=["POST"])
def eliminar_escenario(scenario_id):
    esc = query_one("SELECT * FROM escenarios WHERE id = %s", (scenario_id,))

    if not esc:
        flash("Escenario inexistente.", "error")
        return redirect(url_for("index"))

    if int(esc.get("es_base", 0)) == 1:
        flash("El escenario base no puede eliminarse.", "error")
        return redirect(url_for("index"))

    try:
        execute(
            "UPDATE escenarios SET activo = 0 WHERE id = %s",
            (scenario_id,)
        )
        commit()
        flash("Escenario eliminado.", "ok")
    except Exception as exc:
        rollback()
        flash(f"No se pudo eliminar el escenario: {exc}", "error")

    return redirect(url_for("index"))


@app.route("/escenario/<int:scenario_id>", methods=["GET", "POST"])
def escenario(scenario_id):
    esc = query_one(
        "SELECT * FROM escenarios WHERE id = %s AND activo = 1",
        (scenario_id,)
    )

    if not esc:
        flash("Escenario inexistente o inactivo.", "error")
        return redirect(url_for("index"))

    groups = load_groups()

    if request.method == "POST":
        try:
            defs = query_all("SELECT id, codigo FROM definiciones_parametros")
            param_id_map = {r["codigo"]: r["id"] for r in defs}

            group_id_map = {g["codigo"]: g["id"] for g in groups}

            payload = []

            for param_code in param_id_map:
                for group_code in group_id_map:
                    field = f"{param_code}__{group_code}"
                    raw = request.form.get(field)

                    if raw is None or raw == "":
                        continue

                    value = float(raw.replace(",", "."))
                    payload.append(
                        (
                            scenario_id,
                            param_id_map[param_code],
                            group_id_map[group_code],
                            value
                        )
                    )

            if payload:
                executemany(
                    """
                    INSERT INTO valores_escenario (escenario_id, parametro_id, grupo_id, valor)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE valor = VALUES(valor)
                    """,
                    payload
                )

            base_params = load_base_params()
            _, merged_params = load_parameters_with_values(scenario_id)
            baseline_results = load_baseline_results()

            calc = calculate_results_isolated(
                base_params,
                merged_params,
                baseline_results,
                groups
            )

            persist_results(scenario_id, calc)

            commit()
            flash("Escenario guardado y recalculado.", "ok")

        except Exception as exc:
            rollback()
            flash(f"Error al guardar/recalcular: {exc}", "error")

    modules, _ = load_parameters_with_values(scenario_id)

    result_rows = query_all(
        """
        SELECT d.orden,
               d.descripcion,
               d.codigo,
               g.codigo AS grupo_codigo,
               COALESCE(re.valor, rb.valor) AS valor,
               rb.valor AS base_valor
        FROM definiciones_resultados d
        CROSS JOIN grupos_tarifarios g
        LEFT JOIN resultados_base rb
               ON rb.resultado_id = d.id
              AND rb.grupo_id = g.id
        LEFT JOIN resultados_escenario re
               ON re.resultado_id = d.id
              AND re.grupo_id = g.id
              AND re.escenario_id = %s
        ORDER BY d.orden, g.orden
        """,
        (scenario_id,)
    )

    results = defaultdict(dict)

    for r in result_rows:
        codigo = r["codigo"]
        grupo_codigo = r["grupo_codigo"]

        results[codigo]["label"] = r["descripcion"]
        results[codigo]["order"] = r["orden"]
        results[codigo].setdefault("values", {})
        results[codigo].setdefault("base_values", {})
        results[codigo].setdefault("changed", {})

        value = float(r["valor"]) if r["valor"] is not None else 0.0
        base_value = float(r["base_valor"]) if r["base_valor"] is not None else 0.0

        results[codigo]["values"][grupo_codigo] = value
        results[codigo]["base_values"][grupo_codigo] = base_value
        results[codigo]["changed"][grupo_codigo] = abs(value - base_value) > 1e-12

    ordered_results = sorted(results.items(), key=lambda x: x[1]["order"])
    impact_summary = build_impact_summary(ordered_results, groups, top_n=5)

    return render_template(
        "scenario.html",
        esc=esc,
        groups=groups,
        modules=modules,
        ordered_results=ordered_results,
        impact_summary=impact_summary,
    )

@app.route("/escenario/<int:scenario_id>/exportar_excel")
def exportar_excel(scenario_id):
    esc = query_one(
        "SELECT * FROM escenarios WHERE id=%s AND activo=1",
        (scenario_id,)
    )

    if not esc:
        flash("Escenario inexistente o inactivo.", "error")
        return redirect(url_for("index"))

    grupo = esc.get("grupo_codigo")
    if not grupo:
        flash("El escenario no tiene grupo tarifario asignado.", "error")
        return redirect(url_for("escenario", scenario_id=scenario_id))

    param_rows = query_all(
        """
        SELECT p.descripcion,
               p.codigo,
               b.valor AS valor_base,
               COALESCE(v.valor, b.valor) AS valor_escenario
        FROM definiciones_parametros p
        JOIN grupos_tarifarios g ON g.codigo = %s
        LEFT JOIN valores_base b
               ON b.parametro_id = p.id AND b.grupo_id = g.id
        LEFT JOIN valores_escenario v
               ON v.parametro_id = p.id
              AND v.grupo_id = g.id
              AND v.escenario_id = %s
        ORDER BY p.modulo, p.id
        """,
        (grupo, scenario_id)
    )

    result_rows = query_all(
        """
        SELECT d.descripcion,
               d.codigo,
               rb.valor AS valor_base,
               COALESCE(re.valor, rb.valor) AS valor_escenario
        FROM definiciones_resultados d
        JOIN grupos_tarifarios g ON g.codigo = %s
        LEFT JOIN resultados_base rb
               ON rb.resultado_id = d.id AND rb.grupo_id = g.id
        LEFT JOIN resultados_escenario re
               ON re.resultado_id = d.id
              AND re.grupo_id = g.id
              AND re.escenario_id = %s
        ORDER BY d.orden
        """,
        (grupo, scenario_id)
    )

    wb = Workbook()

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen"

    resumen_map = {r["codigo"]: r for r in result_rows}

    ws["A1"] = "Escenario"
    ws["B1"] = esc["nombre"]
    ws["A2"] = "Grupo tarifario"
    ws["B2"] = grupo

    ws["A4"] = "Costo total por km sin IVA"
    ws["B4"] = float(resumen_map.get("costo_total_km_sin_iva", {}).get("valor_escenario") or 0)

    ws["A5"] = "Tarifa media sin compensación"
    ws["B5"] = float(resumen_map.get("tarifa_media", {}).get("valor_escenario") or 0)

    ws["A6"] = "IPK"
    ws["B6"] = float(resumen_map.get("ipk", {}).get("valor_escenario") or 0)

    for cell in ["A1", "A2", "A4", "A5", "A6"]:
        ws[cell].font = Font(bold=True)

    # Hoja Parámetros
    ws2 = wb.create_sheet("Parametros")
    ws2.append(["Parametro", "Codigo", "Valor base", "Valor escenario", "Delta"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for r in param_rows:
        base = float(r["valor_base"] or 0)
        escv = float(r["valor_escenario"] or 0)
        ws2.append([
            r["descripcion"],
            r["codigo"],
            base,
            escv,
            escv - base
        ])

    # Hoja Resultados
    ws3 = wb.create_sheet("Resultados")
    ws3.append(["Rubro", "Codigo", "Valor base", "Valor escenario", "Delta", "Impacto %"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for r in result_rows:
        base = float(r["valor_base"] or 0)
        escv = float(r["valor_escenario"] or 0)
        delta = escv - base
        pct = (delta / base * 100) if base else 0
        ws3.append([
            r["descripcion"],
            r["codigo"],
            base,
            escv,
            delta,
            pct
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"escenario_{scenario_id}_{grupo}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
if __name__ == "__main__":
    app.run(debug=True)